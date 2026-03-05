import re
from pathlib import Path
import allure
from openpyxl import load_workbook
import pytest
from config import UI_TIMEOUT, STEP_PAUSE
from pages.my_profile_page import MyProfilePage
from pages.sign_in_page import SignInPage
from utils.randoms import random_email

RANDOM_EMAIL_PLACEHOLDER = "{{random_generated_email}}"

def load_update_cases():
    # Read UI update profile test cases from Excel file
    file_path = (
        Path(__file__).resolve().parents[2]
        / "data"
        / "test_cases"
        / "update_profile.xlsx"
    )

    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    if "multiple_fields_ui" in wb.sheetnames:
        ws = wb["multiple_fields_ui"]
    elif "multiple_fields" in wb.sheetnames:
        ws = wb["multiple_fields"]
    else:
        raise KeyError(
            "Worksheet 'multiple_fields_ui' (or legacy 'multiple_fields') does not exist."
        )

    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_values = list(row)
        if len(row_values) < 11:
            row_values.extend([None] * (11 - len(row_values)))

        (
            case_name,
            name,
            email,
            phone,
            old_password,
            new_password,
            confirmation_password,
            is_updated_successful,
            is_updated_fail,
            login_new_success,
            login_old_success,
        ) = row_values[:11]

        if not case_name:
            continue

        case_name = re.sub(r"^\s*\d+\.\s*", "", str(case_name).strip())

        normalized_values = []
        for value in (
            name,
            email,
            phone,
            old_password,
            new_password,
            confirmation_password,
        ):
            if value is None:
                normalized_values.append(None)
                continue

            if isinstance(value, str):
                cleaned = value.strip()
                upper = cleaned.upper()

                if upper == "EMPTY":
                    normalized_values.append("")
                    continue
                if upper == "NULL":
                    normalized_values.append(None)
                    continue

                normalized_values.append(cleaned)
                continue

            normalized_values.append(str(value).strip())

        (
            name,
            email,
            phone,
            old_password,
            new_password,
            confirmation_password,
        ) = normalized_values

        if isinstance(email, str) and email == RANDOM_EMAIL_PLACEHOLDER:
            email = random_email()

        normalized_flags = []
        for value in (
            is_updated_successful,
            is_updated_fail,
            login_new_success,
            login_old_success,
        ):
            if value is None:
                normalized_flags.append(None)
                continue

            if isinstance(value, bool):
                normalized_flags.append(value)
                continue

            if isinstance(value, str):
                lowered = value.strip().lower()
                if lowered == "true":
                    normalized_flags.append(True)
                    continue
                if lowered == "false":
                    normalized_flags.append(False)
                    continue

            normalized_flags.append(bool(value))

        (
            is_updated_successful,
            is_updated_fail,
            login_new_success,
            login_old_success,
        ) = normalized_flags

        cases.append(
            (
                case_name,
                name,
                email,
                phone,
                old_password,
                new_password,
                confirmation_password,
                is_updated_successful,
                is_updated_fail,
                login_new_success,
                login_old_success,
            )
        )

    wb.close()
    return cases


UI_UPDATE_CASES = load_update_cases()


@allure.feature("My Profile UI")
@pytest.mark.parametrize(
    "case_name,name,email,phone,old_password,new_password,confirmation_password,is_updated_successful,is_updated_fail,login_new_success,login_old_success",
    UI_UPDATE_CASES,
    ids=[case[0] for case in UI_UPDATE_CASES],
)
def test_update_profile(
    ui_page,
    account_state,
    case_name,
    name,
    email,
    phone,
    old_password,
    new_password,
    confirmation_password,
    is_updated_successful,
    is_updated_fail,
    login_new_success,
    login_old_success,
):
    # Run one UI profile update case and verify expected result
    page = ui_page

    current_email = account_state.get("email")
    current_password = account_state.get("password")
    if not current_email or not current_password:
        pytest.fail("Missing current account_state email/password")

    page.set_default_timeout(UI_TIMEOUT)
    page.set_default_navigation_timeout(UI_TIMEOUT)
    my_profile_page = MyProfilePage(page)

    allure.dynamic.title(case_name)
    allure.dynamic.description("Update profile directly on the page")

    with allure.step("1. Login"):
        sign_in_page = SignInPage(page)
        sign_in_page.open(timeout=UI_TIMEOUT)
        sign_in_page.is_login_page_loaded(timeout=UI_TIMEOUT)
        page.wait_for_timeout(STEP_PAUSE)
        sign_in_page.login(current_email, current_password)
        page.wait_for_timeout(STEP_PAUSE)

    with allure.step("2. Go to My Profile page."):
        my_profile_page.open_my_profile_page()
        assert my_profile_page.is_my_profile_page_loaded() is True
        page.wait_for_timeout(STEP_PAUSE)

    if case_name == "Verify that user can reload My Profile page and the page still opens normally":
        with allure.step("3. Reload My Profile page"):
            my_profile_page.reload_page()
            page.wait_for_timeout(STEP_PAUSE)

        with allure.step("4. Verify My Profile page is still loaded"):
            assert my_profile_page.is_my_profile_page_loaded() is True
        return

    has_profile_input = any(value is not None for value in (name, email, phone))
    has_password_input = any(
        value is not None for value in (old_password, new_password, confirmation_password)
    )

    if not has_profile_input and not has_password_input:
        return

    if name is not None:
        if name == "":
            with allure.step("3. Empty name"):
                my_profile_page.clear_field(my_profile_page._name_field)
                page.wait_for_timeout(STEP_PAUSE)
        else:
            with allure.step(f'3. Update name to "{name}"'):
                my_profile_page.fill_field(my_profile_page._name_field, name)
                page.wait_for_timeout(STEP_PAUSE)

    if email is not None:
        if email == "":
            with allure.step("3. Empty email"):
                my_profile_page.clear_field(my_profile_page._email_field)
                page.wait_for_timeout(STEP_PAUSE)
        else:
            with allure.step(f'3. Update email to "{email}"'):
                my_profile_page.fill_field(my_profile_page._email_field, email)
                page.wait_for_timeout(STEP_PAUSE)

    if phone is not None:
        if phone == "":
            with allure.step("3. Empty phone"):
                my_profile_page.clear_field(my_profile_page._phone_field)
                page.wait_for_timeout(STEP_PAUSE)
        else:
            with allure.step(f'3. Update phone to "{phone}"'):
                my_profile_page.fill_field(my_profile_page._phone_field, phone)
                page.wait_for_timeout(STEP_PAUSE)

    if old_password is not None:
        if old_password == "":
            with allure.step("3. Empty old_password"):
                my_profile_page.clear_field(my_profile_page._old_password_field)
                page.wait_for_timeout(STEP_PAUSE)
        else:
            with allure.step('3. Update old_password to "***"'):
                my_profile_page.fill_field(my_profile_page._old_password_field, old_password)
                page.wait_for_timeout(STEP_PAUSE)

    if new_password is not None:
        if new_password == "":
            with allure.step("3. Empty new_password"):
                my_profile_page.clear_field(my_profile_page._new_password_field)
                page.wait_for_timeout(STEP_PAUSE)
        else:
            with allure.step('3. Update new_password to "***"'):
                my_profile_page.fill_field(my_profile_page._new_password_field, new_password)
                page.wait_for_timeout(STEP_PAUSE)

    if confirmation_password is not None:
        if confirmation_password == "":
            with allure.step("3. Empty confirmation_password"):
                my_profile_page.clear_field(my_profile_page._confirm_password_field)
                page.wait_for_timeout(STEP_PAUSE)
        else:
            with allure.step('3. Update confirmation_password to "***"'):
                my_profile_page.fill_field(my_profile_page._confirm_password_field, confirmation_password)
                page.wait_for_timeout(STEP_PAUSE)

    require_enabled_save_button = bool(is_updated_successful) or bool(login_new_success)

    with allure.step("4. Click Save Profile button"):
        my_profile_page.click_save_profile(require_enabled=require_enabled_save_button)
        page.wait_for_timeout(STEP_PAUSE)

    expected_name = name if name is not None else None
    expected_email = email if email is not None else None
    expected_phone = phone if phone is not None else None

    is_old_password_missing = (old_password in (None, "")) and (
        new_password is not None or confirmation_password is not None
    )

    failure_expected = bool(is_updated_fail) or (has_password_input and is_old_password_missing)

    if login_new_success:
        if not new_password:
            pytest.fail(f"[{case_name}] login_new_success is TRUE but new_password is empty.")

        assert my_profile_page.is_updated_successful() is True

        with allure.step("5. Log out"):
            my_profile_page.user_avatar.log_out()
            page.wait_for_timeout(STEP_PAUSE)

        with allure.step("6. Login again"):
            sign_in_page = SignInPage(page)
            sign_in_page.is_login_page_loaded(timeout=UI_TIMEOUT)
            login_email = current_email
            if isinstance(email, str) and email.strip():
                login_email = email
            sign_in_page.login(login_email, new_password)
            page.wait_for_timeout(STEP_PAUSE)

        verify_profile_page = MyProfilePage(page)
        verify_profile_page.open_my_profile_page()
        assert verify_profile_page.is_my_profile_page_loaded() is True

        if isinstance(login_email, str) and login_email.strip():
            account_state["email"] = login_email
        account_state["password"] = new_password
        return

    if is_updated_successful:
        assert my_profile_page.is_updated_successful(
            expected_name=expected_name,
            expected_email=expected_email,
            expected_phone=expected_phone,
        ) is True
        if isinstance(email, str) and email.strip():
            account_state["email"] = email
        if isinstance(new_password, str) and new_password.strip():
            account_state["password"] = new_password
        return

    if failure_expected:
        assert my_profile_page.is_updated_fail() is True

    if login_old_success:
        with allure.step("5. Log out"):
            my_profile_page.user_avatar.log_out()
            page.wait_for_timeout(STEP_PAUSE)

        with allure.step("6. Login again"):
            sign_in_page = SignInPage(page)
            sign_in_page.is_login_page_loaded(timeout=UI_TIMEOUT)
            sign_in_page.login(account_state["email"], account_state["password"])
            page.wait_for_timeout(STEP_PAUSE)

        verify_profile_page = MyProfilePage(page)
        verify_profile_page.open_my_profile_page()
        assert verify_profile_page.is_my_profile_page_loaded() is True
        return

    if failure_expected:
        return

    pytest.skip("No expected result provided in this row.")
