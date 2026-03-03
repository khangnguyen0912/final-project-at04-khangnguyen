import json
from pathlib import Path
import allure
from openpyxl import load_workbook
import pytest
from api.api_list import APIs
from api.core.endpoints import LOGIN
from config import PASSWORD
from helper.assertions import (
    assert_status_code,
    assert_success_message,
    assert_validation_error,
)

pw_timeout = 3000


def load_password_cases():
    file_path = (
        Path(__file__).resolve().parents[4]
        / "data"
        / "test_cases"
        / "update_profile.xlsx"
    )
    wb = load_workbook(file_path, data_only=True)
    ws = wb["password"]

    cases = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_values = list(row)
        if len(row_values) < 8:
            row_values.extend([None] * (8 - len(row_values)))

        (
            case_name,
            password_old,
            password,
            expected_status,
            success_message,
            login_old_success,
            login_new_success,
            error_msg,
        ) = row_values[:8]

        if not case_name:
            continue

        normalized_password_values = []
        for value in (password_old, password):
            if isinstance(value, str):
                upper = value.upper()
                if upper == "NULL":
                    normalized_password_values.append(None)
                    continue
                if upper == "EMPTY":
                    normalized_password_values.append("")
                    continue
            normalized_password_values.append(value)

        password_old, password = normalized_password_values

        normalized_login_flags = []
        for value in (login_old_success, login_new_success):
            if value is None:
                normalized_login_flags.append(None)
                continue
            if isinstance(value, bool):
                normalized_login_flags.append(value)
                continue
            if isinstance(value, str):
                lowered = value.strip().lower()
                if lowered == "true":
                    normalized_login_flags.append(True)
                    continue
                if lowered == "false":
                    normalized_login_flags.append(False)
                    continue
            normalized_login_flags.append(bool(value))

        login_old_success, login_new_success = normalized_login_flags

        if expected_status is None or str(expected_status).strip() == "":
            raise ValueError(
                f"Row {row_idx}: expected_status is required for case '{case_name}'."
            )

        try:
            expected_status_int = int(expected_status)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Row {row_idx}: expected_status must be an int for case '{case_name}', got '{expected_status}'."
            ) from exc

        cases.append(
            (
                str(case_name),
                password_old,
                password,
                expected_status_int,
                success_message,
                login_old_success,
                login_new_success,
                error_msg,
            )
        )

    wb.close()
    return cases


PASSWORD_CASES = load_password_cases()


async def _try_get_access_token(api_request, email: str, password):
    if password is None:
        return None

    payload = {"email": email, "password": password}
    res = await api_request.post(
        LOGIN,
        data=json.dumps(payload),
        timeout=pw_timeout,
    )

    try:
        body = await res.json()
    except Exception:
        return None

    token = body.get("accessToken")
    if isinstance(token, str) and token.strip():
        return token
    return None


def _resolve_password_old(password_old, account_state):
    current_password = account_state.get("password")
    if (
        isinstance(password_old, str)
        and isinstance(current_password, str)
        and password_old == PASSWORD
        and current_password.strip()
    ):
        # Keep "valid old password" cases stable after previous password changes.
        return current_password
    return password_old


async def _assert_login_result(api_request, email: str, password, expected_success: bool, case_name: str):
    token = await _try_get_access_token(api_request, email, password)
    actual_success = bool(token)
    assert actual_success == expected_success, (
        f"[{case_name}] Expected login_success={expected_success} with provided password, "
        f"but got {actual_success}."
    )


@allure.feature("Profile API")
@pytest.mark.asyncio
@pytest.mark.parametrize(
    "case_name,password_old,password,expected_status,success_message,login_old_success,login_new_success,error_msg",
    PASSWORD_CASES,
    ids=[case[0] for case in PASSWORD_CASES],
)
async def test_update_profile_with_password_field_only(
    api_request,
    account_state,
    case_name,
    password_old,
    password,
    expected_status,
    success_message,
    login_old_success,
    login_new_success,
    error_msg,
):
    current_email = account_state.get("email")
    current_password = account_state.get("password")
    if not current_email or not current_password:
        pytest.fail("Missing current account_state email/password")

    effective_password_old = _resolve_password_old(password_old, account_state)

    allure.dynamic.title(case_name)
    allure.dynamic.description(
        "Send PATCH /api/profile to update password only, then verify login behavior."
    )

    with allure.step("1. Prepare authorization token"):
        token = await _try_get_access_token(api_request, current_email, current_password)
        assert token, f"[{case_name}] Cannot login with current account_state credentials."
        api = APIs(api_request, token)

    with allure.step("2. Send PATCH /api/profile (update password)"):
        payload = {"password_old": effective_password_old, "password": password}
        res = await api.update_profile(json.dumps(payload), timeout=pw_timeout)

    with allure.step("3. Verify the response"):
        with allure.step(f"3.1 Verify status code is {expected_status}"):
            await assert_status_code(res, expected_status, case_name)

        if expected_status == 200:
            expected_success_message = success_message or "Updated profile successfully."
            with allure.step("3.2 Verify success message"):
                await assert_success_message(res, str(expected_success_message))
        elif error_msg:
            with allure.step("3.2 Verify error validation message"):
                await assert_validation_error(res, "password", str(error_msg))
            with allure.step("3.3 Attach response body"):
                try:
                    error_body = await res.text()
                except Exception as exc:
                    error_body = f"<cannot read response text: {exc}>"
                allure.attach(
                    error_body,
                    "Validation error response",
                    allure.attachment_type.TEXT,
                )

    with allure.step("4. Verify functionality of Login API"):
        if login_old_success is not None:
            with allure.step("4.1 Verify login by old password"):
                await _assert_login_result(
                    api_request,
                    current_email,
                    effective_password_old,
                    bool(login_old_success),
                    case_name,
                )

        if login_new_success is not None:
            with allure.step("4.2 Verify login by new password"):
                await _assert_login_result(
                    api_request,
                    current_email,
                    password,
                    bool(login_new_success),
                    case_name,
                )

    if expected_status == 200 and bool(login_new_success) and isinstance(password, str) and password.strip():
        account_state["password"] = password
