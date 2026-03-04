import json
from pathlib import Path
import allure
from openpyxl import load_workbook
import pytest
from api.api_list import APIs
from helper.assertions import (
    assert_status_code,
    assert_success_message,
    assert_field_updated,
    assert_other_fields_unchanged,
    assert_validation_error,
)
from utils.randoms import random_email

pw_timeout = 3000
RANDOM_EMAIL_PLACEHOLDER = "{{random_generated_email}}"

def load_email_cases():
    # Read email test cases from Excel file
    file_path = (
        Path(__file__).resolve().parents[4]
        / "data"
        / "test_cases"
        / "update_profile.xlsx"
    )
    wb = load_workbook(file_path, data_only=True)
    ws = wb["email"]

    cases = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        (
            case_name,
            email,
            expected_status,
            _success_message,
            expected_email_after,
            verify_others,
            error_msg,
        ) = row

        if not case_name:
            continue

        if isinstance(email, str) and email.upper() == "NULL":
            email = None

        if isinstance(email, str) and email.upper() == "EMPTY":
            email = ""

        if isinstance(email, str) and email.strip() == RANDOM_EMAIL_PLACEHOLDER:
            email = random_email()
            expected_email_after = email

        if (
            isinstance(expected_email_after, str)
            and expected_email_after.strip() == RANDOM_EMAIL_PLACEHOLDER
        ):
            expected_email_after = email

        if isinstance(verify_others, str):
            verify_others = verify_others.strip().lower() == "true"

        if expected_status is None or str(expected_status).strip() == "":
            raise ValueError(
                f"Row {row_idx}: expected_status is required for case '{case_name}'."
            )

        try:
            expected_status_int = int(expected_status)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Row {row_idx}: expected_status must be an int for case '{case_name}', got '{expected_status}'."
            ) from exc

        cases.append(
            (
                str(case_name),
                email,
                expected_status_int,
                expected_email_after,
                bool(verify_others),
                error_msg,
            )
        )

    return cases


EMAIL_CASES = load_email_cases()


@allure.feature("Profile API")
@pytest.mark.asyncio
@pytest.mark.parametrize(
    "case_name,email,expected_status,expected_email_after,verify_others,error_msg",
    EMAIL_CASES,
    ids=[case[0] for case in EMAIL_CASES],
)
async def test_update_profile_with_email_field_only(
    api_request,
    access_token,
    account_state,
    case_name,
    email,
    expected_status,
    expected_email_after,
    verify_others,
    error_msg,
):
    # Update email and verify API response and profile data
    api = APIs(api_request, access_token)

    allure.dynamic.title(case_name)
    allure.dynamic.description(
        "Send PATCH /api/profile to update email only, then verify using GET /api/me."
    )

    before = None
    with allure.step("1. Get current user information"):
        if expected_status == 200 and verify_others:
            with allure.step("1.1 Send GET /api/me (baseline)"):
                res_before = await api.get_me(timeout=pw_timeout)
                assert res_before.ok, (
                    f"[{case_name}] GET /api/me failed | {res_before.status} | {await res_before.text()}"
                )
                before = await res_before.json()

    with allure.step("2. Send PATCH /api/profile (update email)"):
        payload = {"email": email}
        res = await api.update_profile(json.dumps(payload), timeout=pw_timeout)

    with allure.step("3. Verify the response"):
        with allure.step(f"3.1 Verify status code is {expected_status}"):
            await assert_status_code(res, expected_status, case_name)

        if expected_status == 200:
            with allure.step("3.2 Verify success message"):
                await assert_success_message(res, "Updated profile successfully.")
        elif error_msg:
            with allure.step("3.2 Verify error validation message"):
                await assert_validation_error(res, "email", str(error_msg))
            with allure.step("3.3 Attach response body"):
                try:
                    error_body = await res.text()
                except Exception as exc:
                    error_body = f"<cannot read response text: {exc}>"
                allure.attach(
                    error_body,
                    "Validation error response",
                    allure.attachment_type.TEXT,
                )

    if expected_status == 200:
        with allure.step("4. Verify profile data after update"):
            with allure.step("4.1 Send GET /api/me (after update)"):
                res_after = await api.get_me(timeout=pw_timeout)
                assert res_after.ok, (
                    f"[{case_name}] GET /api/me failed | {res_after.status} | {await res_after.text()}"
                )
                after = await res_after.json()

            with allure.step("4.2 Verify updated field: email"):
                assert_field_updated(after, "email", expected_email_after)

            if isinstance(expected_email_after, str) and expected_email_after.strip():
                account_state["email"] = expected_email_after

            if verify_others:
                with allure.step("4.3 Verify other fields unchanged"):
                    assert_other_fields_unchanged(after, before, ["name", "phone"])
